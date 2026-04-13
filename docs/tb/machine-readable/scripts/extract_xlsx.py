#!/usr/bin/env python3
"""
Extract schema, sample data, validation rules, and macro inventory
from TB/PL Excel workbooks (.xlsm).

Uses openpyxl in read_only mode for data scanning (critical for 74MB+ files),
and full mode for formula/validation/named-range extraction.
"""

import csv
import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import yaml

TB_DIR = Path(__file__).resolve().parent.parent.parent  # docs/tb/
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "training-data"

WORKBOOKS = [
    {
        "filename": "HKG_TB review Nov'25.xlsm",
        "id": "hkg-tb",
        "name": "HKG TB Review Nov 2025",
        "schema_output": "hkg-tb-schema.yaml",
        "sample_output": "hkg-tb-sample.csv",
        "validation_output": "hkg-tb-validation-rules.yaml",
        "named_ranges_output": "hkg-tb-named-ranges.yaml",
    },
    {
        "filename": "HKG_PL review Nov'25.xlsm",
        "id": "hkg-pl",
        "name": "HKG PL Review Nov 2025",
        "schema_output": "hkg-pl-schema.yaml",
        "sample_output": "hkg-pl-sample.csv",
        "validation_output": "hkg-pl-validation-rules.yaml",
        "named_ranges_output": None,  # share the TB one
    },
]

# Type inference patterns
MEASURE_PATTERNS = [
    "amount", "balance", "total", "sum", "variance", "debit", "credit",
    "net", "gross", "value", "mtm", "exposure", "pv", "ytd", "mtd",
    "actual", "budget", "forecast", "prior", "current",
]
DATE_PATTERNS = ["date", "period", "month", "year", "quarter"]
ID_PATTERNS = ["_id", "account", "code", "gl_", "cost_center", "entity"]


def infer_data_type(values: list) -> str:
    """Infer column data type from sample values."""
    non_null = [v for v in values if v is not None]
    if not non_null:
        return "UNKNOWN"

    type_counts = {"int": 0, "float": 0, "str": 0, "date": 0, "bool": 0}
    for v in non_null[:50]:
        if isinstance(v, bool):
            type_counts["bool"] += 1
        elif isinstance(v, int):
            type_counts["int"] += 1
        elif isinstance(v, float):
            type_counts["float"] += 1
        elif isinstance(v, datetime):
            type_counts["date"] += 1
        else:
            type_counts["str"] += 1

    dominant = max(type_counts, key=type_counts.get)
    return {
        "int": "INTEGER",
        "float": "DECIMAL",
        "str": "NVARCHAR",
        "date": "DATE",
        "bool": "BOOLEAN",
    }.get(dominant, "NVARCHAR")


def classify_field(col_name: str) -> Dict[str, Optional[str]]:
    """Classify field type using naming heuristics."""
    name_lower = col_name.lower()

    for p in ID_PATTERNS:
        if p in name_lower:
            return {"fieldType": "identifier", "analyticsAnnotation": None}

    for p in DATE_PATTERNS:
        if p in name_lower:
            return {"fieldType": "date", "analyticsAnnotation": None}

    for p in MEASURE_PATTERNS:
        if p in name_lower:
            return {"fieldType": "measure", "analyticsAnnotation": "@Analytics.Measure"}

    return {"fieldType": "dimension", "analyticsAnnotation": "@Analytics.Dimension"}


def extract_sheet_schema(ws, sheet_name: str, max_sample_rows: int = 50) -> Dict:
    """Extract schema from a single worksheet."""
    rows_iter = ws.iter_rows(values_only=True)

    # Find header row (first non-empty row)
    header_row = None
    header_row_idx = 0
    for i, row in enumerate(rows_iter):
        non_empty = [c for c in row if c is not None]
        if len(non_empty) >= 2:  # at least 2 non-empty cells = likely header
            header_row = row
            header_row_idx = i + 1
            break

    if header_row is None:
        return {"sheetName": sheet_name, "empty": True, "fields": []}

    headers = []
    for j, h in enumerate(header_row):
        if h is not None:
            headers.append((j, str(h).strip()))

    if not headers:
        return {"sheetName": sheet_name, "empty": True, "fields": []}

    # Read sample data rows
    sample_rows = []
    for row in rows_iter:
        if len(sample_rows) >= max_sample_rows:
            break
        sample_rows.append(row)

    # Build field schemas
    fields = []
    for col_idx, col_name in headers:
        values = []
        for row in sample_rows:
            if col_idx < len(row):
                values.append(row[col_idx])

        non_null = [v for v in values if v is not None]
        null_count = len(values) - len(non_null)
        classification = classify_field(col_name)

        col_letter = openpyxl.utils.get_column_letter(col_idx + 1)

        field = {
            "technicalName": re.sub(r'[^a-zA-Z0-9_]', '_', col_name.upper()).strip('_'),
            "columnLetter": col_letter,
            "columnIndex": col_idx,
            "businessName": col_name,
            "fieldType": classification["fieldType"],
            "dataType": infer_data_type(values),
            "analyticsAnnotation": classification["analyticsAnnotation"],
            "nullRate": round(null_count / max(len(values), 1), 3),
            "sampleValues": [str(v)[:80] for v in non_null[:5]],
        }
        fields.append(field)

    return {
        "sheetName": sheet_name,
        "headerRow": header_row_idx,
        "dataRowCount": len(sample_rows),
        "fieldCount": len(fields),
        "fields": fields,
    }


def extract_formulas_and_validations(wb_full, sheet_name: str) -> List[Dict]:
    """Extract formulas and data validations from a sheet (requires full-mode workbook)."""
    rules = []

    try:
        ws = wb_full[sheet_name]
    except KeyError:
        return rules

    rule_id = 0

    # Scan for formulas in first 200 rows
    for row in ws.iter_rows(min_row=1, max_row=min(200, ws.max_row or 1), values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                rule_id += 1
                formula = cell.value
                # Classify formula type
                formula_type = "formula"
                if "VLOOKUP" in formula.upper():
                    formula_type = "vlookup"
                elif "SUMIF" in formula.upper():
                    formula_type = "sumif"
                elif "IF(" in formula.upper():
                    formula_type = "conditional"
                elif "INDEX" in formula.upper() or "MATCH" in formula.upper():
                    formula_type = "index_match"

                rules.append({
                    "ruleId": f"VR-{rule_id:03d}",
                    "sheet": sheet_name,
                    "cell": cell.coordinate,
                    "type": formula_type,
                    "formula": formula[:200],
                })

                if rule_id >= 500:  # cap to avoid massive output
                    break
        if rule_id >= 500:
            break

    # Data validations
    if hasattr(ws, 'data_validations') and ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            rule_id += 1
            rules.append({
                "ruleId": f"VR-{rule_id:03d}",
                "sheet": sheet_name,
                "type": "data_validation",
                "validationType": dv.type or "unknown",
                "formula1": str(dv.formula1)[:200] if dv.formula1 else None,
                "formula2": str(dv.formula2)[:200] if dv.formula2 else None,
                "sqref": str(dv.sqref)[:100] if dv.sqref else None,
                "allowBlank": dv.allow_blank,
            })

    # Conditional formatting
    if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                rule_id += 1
                rules.append({
                    "ruleId": f"VR-{rule_id:03d}",
                    "sheet": sheet_name,
                    "type": "conditional_format",
                    "priority": rule.priority,
                    "ruleType": rule.type,
                    "formula": str(rule.formula)[:200] if hasattr(rule, 'formula') and rule.formula else None,
                    "sqref": str(cf.sqref)[:100] if hasattr(cf, 'sqref') else None,
                })

    return rules


def extract_named_ranges(wb_full) -> List[Dict]:
    """Extract named ranges from workbook."""
    ranges = []
    for name in wb_full.defined_names.definedName:
        try:
            ranges.append({
                "name": name.name,
                "refersTo": str(name.attr_text)[:200],
                "scope": "workbook" if name.localSheetId is None else f"sheet_{name.localSheetId}",
                "hidden": getattr(name, 'hidden', False),
            })
        except Exception:
            pass
    return ranges


def extract_macro_inventory(filepath: Path) -> Dict:
    """Extract VBA macro module names and function signatures."""
    try:
        from oletools.olevba import VBA_Parser
        vba_parser = VBA_Parser(str(filepath))

        if not vba_parser.detect_vba_macros():
            return {"hasMacros": False}

        modules = []
        public_functions = []
        total_lines = 0

        for _, _, vba_filename, vba_code in vba_parser.extract_macros():
            lines = vba_code.split("\n")
            total_lines += len(lines)
            modules.append(vba_filename)

            # Extract function/sub signatures (not full code)
            for line in lines:
                stripped = line.strip()
                if re.match(r'(Public |Private )?(Sub|Function)\s+\w+', stripped):
                    # Extract just the signature
                    sig = stripped.split("'")[0].strip()  # remove inline comments
                    public_functions.append(sig[:150])

        vba_parser.close()
        return {
            "hasMacros": True,
            "moduleCount": len(modules),
            "modules": modules,
            "functionCount": len(public_functions),
            "publicFunctions": public_functions[:50],  # cap
            "totalCodeLines": total_lines,
        }
    except ImportError:
        return {"hasMacros": "unknown", "note": "oletools not installed"}
    except Exception as e:
        return {"hasMacros": "error", "error": str(e)[:200]}


def extract_sample_csv(ws, output_path: Path, max_rows: int = 100):
    """Extract first N rows as CSV."""
    rows_written = 0
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            if rows_written > max_rows:
                break
            writer.writerow([str(c)[:200] if c is not None else "" for c in row])
            rows_written += 1
    return rows_written


def process_workbook(wb_config: Dict):
    """Process a single workbook end-to-end."""
    filepath = TB_DIR / wb_config["filename"]
    if not filepath.exists():
        print(f"[SKIP] {wb_config['filename']} not found")
        return

    file_size = filepath.stat().st_size
    print(f"\n{'='*60}")
    print(f"Processing: {wb_config['filename']} ({file_size / 1024 / 1024:.1f} MB)")
    print(f"{'='*60}")

    # Phase 1: Read-only mode for sheet inventory and schemas
    print("\n  Phase 1: Schema extraction (read-only mode)...")
    wb_ro = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)

    sheet_schemas = []
    for sheet_name in wb_ro.sheetnames:
        ws = wb_ro[sheet_name]
        schema = extract_sheet_schema(ws, sheet_name)
        sheet_schemas.append(schema)
        field_count = schema.get("fieldCount", 0)
        if field_count > 0:
            print(f"    Sheet '{sheet_name}': {field_count} fields")

    # Extract sample CSV from first significant sheet
    print("\n  Phase 2: Sample data extraction...")
    sample_path = OUTPUT_DIR / wb_config["sample_output"]
    for sheet_name in wb_ro.sheetnames:
        ws = wb_ro[sheet_name]
        # Find first sheet with data
        first_row = next(ws.iter_rows(values_only=True, max_row=1), None)
        if first_row and any(c is not None for c in first_row):
            rows_written = extract_sample_csv(ws, sample_path)
            print(f"    -> {sample_path} ({rows_written} rows from '{sheet_name}')")
            break

    wb_ro.close()

    # Phase 3: Full mode for formulas, validations, named ranges
    print("\n  Phase 3: Formula/validation extraction (full mode)...")
    try:
        wb_full = openpyxl.load_workbook(str(filepath), read_only=False, data_only=False)

        all_rules = []
        for sheet_name in wb_full.sheetnames:
            rules = extract_formulas_and_validations(wb_full, sheet_name)
            all_rules.extend(rules)
            if rules:
                print(f"    Sheet '{sheet_name}': {len(rules)} rules")

        # Named ranges
        named_ranges = extract_named_ranges(wb_full)
        print(f"    Named ranges: {len(named_ranges)}")

        wb_full.close()
    except Exception as e:
        print(f"    [WARN] Full-mode load failed ({e}), skipping formulas")
        all_rules = []
        named_ranges = []

    # Phase 4: VBA macro inventory
    print("\n  Phase 4: VBA macro inventory...")
    macro_info = extract_macro_inventory(filepath)
    if macro_info.get("hasMacros"):
        print(f"    Modules: {macro_info.get('moduleCount', '?')}")
        print(f"    Functions: {macro_info.get('functionCount', '?')}")

    # Write schema YAML
    schema_doc = {
        "workbookId": wb_config["id"],
        "name": wb_config["name"],
        "sourceFile": wb_config["filename"],
        "extractedAt": datetime.now(timezone.utc).isoformat(),
        "fileSize": file_size,
        "sheetCount": len(sheet_schemas),
        "sheets": sheet_schemas,
        "macroInventory": macro_info,
    }
    schema_path = OUTPUT_DIR / wb_config["schema_output"]
    with open(schema_path, "w", encoding="utf-8") as f:
        yaml.dump(schema_doc, f, default_flow_style=False, sort_keys=False, allow_unicode=True, width=120)
    print(f"\n  -> {schema_path}")

    # Write validation rules YAML
    if all_rules:
        val_doc = {
            "workbookId": wb_config["id"],
            "sourceFile": wb_config["filename"],
            "extractedAt": datetime.now(timezone.utc).isoformat(),
            "ruleCount": len(all_rules),
            "validationRules": all_rules,
        }
        val_path = OUTPUT_DIR / wb_config["validation_output"]
        with open(val_path, "w", encoding="utf-8") as f:
            yaml.dump(val_doc, f, default_flow_style=False, sort_keys=False, allow_unicode=True, width=120)
        print(f"  -> {val_path}")

    # Write named ranges YAML
    if named_ranges and wb_config.get("named_ranges_output"):
        nr_doc = {
            "workbookId": wb_config["id"],
            "sourceFile": wb_config["filename"],
            "extractedAt": datetime.now(timezone.utc).isoformat(),
            "rangeCount": len(named_ranges),
            "namedRanges": named_ranges,
        }
        nr_path = OUTPUT_DIR / wb_config["named_ranges_output"]
        with open(nr_path, "w", encoding="utf-8") as f:
            yaml.dump(nr_doc, f, default_flow_style=False, sort_keys=False, allow_unicode=True, width=120)
        print(f"  -> {nr_path}")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("Excel Workbook Extraction Pipeline")
    print("=" * 60)

    for wb_config in WORKBOOKS:
        process_workbook(wb_config)

    print(f"\n{'='*60}")
    print(f"Done. Output in {OUTPUT_DIR}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
