#!/usr/bin/env python3
"""Tests for the Excel-to-CSV pre-converter."""

import csv
import os
import tempfile

import pytest


def test_sanitise_name():
    from excel_to_csv import sanitise_name

    assert sanitise_name("Sheet 1") == "Sheet_1"
    assert sanitise_name("Treasury/ESG") == "Treasury_ESG"
    assert sanitise_name("OK_name-2.csv") == "OK_name-2.csv"


def test_convert_workbook_creates_csvs(tmp_path):
    """Requires openpyxl — creates a minimal workbook in memory."""
    from openpyxl import Workbook
    from excel_to_csv import convert_workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws.append(["col1", "col2"])
    ws.append([1, "hello"])
    ws.append([2, "world"])

    xlsx_path = str(tmp_path / "test.xlsx")
    wb.save(xlsx_path)
    wb.close()

    out_dir = str(tmp_path / "csv_out")
    paths = convert_workbook(xlsx_path, out_dir)

    assert len(paths) == 1
    assert os.path.exists(paths[0])

    with open(paths[0]) as f:
        rows = list(csv.reader(f))
    assert len(rows) == 3
    assert rows[0] == ["col1", "col2"]
    assert rows[1] == ["1", "hello"]


def test_convert_directory_skips_temp_files(tmp_path):
    """Files with ~ prefix or [nn] suffix should be skipped."""
    from openpyxl import Workbook
    from excel_to_csv import convert_directory

    # Create a normal file
    wb = Workbook()
    wb.active.append(["a"])
    wb.save(str(tmp_path / "good.xlsx"))
    wb.close()

    # Create files that should be skipped
    for name in ["~$lock.xlsx", "data[95].xlsx"]:
        wb2 = Workbook()
        wb2.save(str(tmp_path / name))
        wb2.close()

    paths = convert_directory(str(tmp_path), str(tmp_path / "out"))
    assert len(paths) == 1  # only "good.xlsx"

