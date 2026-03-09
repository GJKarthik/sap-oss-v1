#!/usr/bin/env python3
"""
Excel-to-CSV pre-converter for the text-to-SQL pipeline.

Converts all .xlsx files in a data directory to per-sheet CSV files
so downstream Zig stages can work with plain text.

Usage:
    python3 excel_to_csv.py ../../data ../../pipeline/output/intermediate/csv
"""

import os
import sys
import csv
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # Handled at call sites; allows module to be imported for tests


def sanitise_name(name: str) -> str:
    """Sanitise a filename to be safe across platforms."""
    return "".join(c if c.isalnum() or c in "-_." else "_" for c in name)


def convert_workbook(xlsx_path: str, output_dir: str) -> list[str]:
    """Convert all sheets in an .xlsx file to separate CSV files.

    Returns a list of output CSV file paths.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    stem = Path(xlsx_path).stem
    os.makedirs(output_dir, exist_ok=True)
    paths: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_sheet = sanitise_name(sheet_name)
        out_path = os.path.join(output_dir, f"{stem}__{safe_sheet}.csv")
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([str(cell) if cell is not None else "" for cell in row])
        paths.append(out_path)

    wb.close()
    return paths


def convert_directory(data_dir: str, output_dir: str) -> list[str]:
    """Convert every .xlsx file in *data_dir*."""
    all_paths: list[str] = []
    data = Path(data_dir)
    for xlsx in sorted(data.glob("*.xlsx")):
        # Skip temp/lock files
        if xlsx.name.startswith("~") or "[" in xlsx.name:
            continue
        print(f"  Converting {xlsx.name} …")
        all_paths.extend(convert_workbook(str(xlsx), output_dir))
    return all_paths


if __name__ == "__main__":
    data_dir = sys.argv[1] if len(sys.argv) > 1 else "../../data"
    outdir = sys.argv[2] if len(sys.argv) > 2 else "../../pipeline/output/intermediate/csv"
    print(f"Pre-converting xlsx files from {data_dir} → {outdir}")
    paths = convert_directory(data_dir, outdir)
    print(f"Done — {len(paths)} CSV files written")
    for p in paths:
        print(f"  {p}")

