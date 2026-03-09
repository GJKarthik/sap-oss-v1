#!/usr/bin/env python3
"""
XLSX-to-ODPS 4.1 Conversion & Enrichment Pipeline

Reads HANA metadata from training-main/data/ XLSX files, profiles them
using data-cleaning-copilot utilities, and outputs enriched ODPS 4.1
YAML with quality annotations and missing-data flags.

Integration points:
  - data-cleaning-copilot: profiling (util_profiler), column mapping
  - ai-core-pal: output wired to PAL input ports
  - Mangle: field classification rules
"""

import sys
import os
import json
import yaml
import csv
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime

# Add data-cleaning-copilot to path for profiler and column mapping utilities
COPILOT_PATH = Path(__file__).resolve().parent.parent.parent / "sap-oss" / "data-cleaning-copilot-main"
sys.path.insert(0, str(COPILOT_PATH))

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data_products" / "enriched"


def try_import_profiler():
    """Try to import data-cleaning-copilot profiler; graceful fallback."""
    try:
        from definition.base.util_profiler import profile_table_data, profile_table_column_data
        return profile_table_data, profile_table_column_data
    except ImportError:
        print("[WARN] data-cleaning-copilot profiler not available, using basic profiling")
        return None, None


def try_import_column_mapping():
    """Try to import data-cleaning-copilot column mapping utilities."""
    try:
        from definition.base.util_column_mapping import (
            normalize_column_name, create_column_mapping, find_matching_table_column
        )
        return normalize_column_name, create_column_mapping, find_matching_table_column
    except ImportError:
        print("[WARN] data-cleaning-copilot column mapping not available")
        return None, None, None


def load_xlsx_sheet(filepath: str, sheet_name: str = None) -> Tuple[List[str], List[List]]:
    """Load an XLSX sheet, return (headers, rows)."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = [[c for c in row] for row in rows[1:]]
    return headers, data


def basic_profile(headers: List[str], data: List[List]) -> Dict[str, Any]:
    """Basic profiling without pandas/ydata dependency."""
    profile = {
        "row_count": len(data),
        "column_count": len(headers),
        "columns": {}
    }
    for i, col_name in enumerate(headers):
        values = [row[i] for row in data if i < len(row) and row[i] is not None]
        non_null = len(values)
        unique = len(set(str(v) for v in values))
        profile["columns"][col_name] = {
            "non_null_count": non_null,
            "null_count": len(data) - non_null,
            "completeness_pct": round(non_null / max(len(data), 1) * 100, 1),
            "unique_count": unique,
            "sample_values": [str(v)[:60] for v in values[:5]]
        }
    return profile


def pandas_profile(headers, data, profile_fn):
    """Full profiling using data-cleaning-copilot's util_profiler."""
    import pandas as pd
    df = pd.DataFrame(data, columns=headers)
    return profile_fn(df, max_columns=len(headers), sample_size=5)


def classify_field(col_name: str, values: List, normalize_fn=None) -> Dict[str, str]:
    """
    Classify a field as dimension/measure/identifier/date using Mangle-style rules.
    Mirrors the classification logic in mangle/domain/hana_fields.mg.
    """
    name_lower = col_name.lower()

    # Identifier patterns
    id_patterns = ["_id", "cusip", "isin", "belnr", "vbeln", "kunnr", "lifnr"]
    if any(p in name_lower for p in id_patterns):
        return {"fieldType": "identifier", "analyticsAnnotation": None}

    # Date patterns
    date_patterns = ["date", "period", "calmonth", "month", "year", "gjahr"]
    if any(p in name_lower for p in date_patterns):
        return {"fieldType": "date", "analyticsAnnotation": None}

    # Measure patterns
    measure_patterns = [
        "amount", "value", "total", "sum", "avg", "count", "usd", "mtm",
        "notional", "rwa", "pv01", "delta", "yield", "price", "revenue",
        "cost", "emission", "intensity", "exposure", "score", "assets",
        "production", "factor"
    ]
    if any(p in name_lower for p in measure_patterns):
        return {"fieldType": "measure", "analyticsAnnotation": "@Analytics.Measure"}

    # Currency patterns
    ccy_patterns = ["ccy", "curr", "waerk", "rhcur", "currency"]
    if any(p in name_lower for p in ccy_patterns):
        return {"fieldType": "currency", "analyticsAnnotation": None}

    # Default: dimension
    return {"fieldType": "dimension", "analyticsAnnotation": "@Analytics.Dimension"}


def detect_missing_data(profile: Dict) -> List[Dict]:
    """Detect fields with missing or incomplete metadata."""
    issues = []
    columns = profile.get("columns", {})
    if not columns:
        # Handle pandas profiler output format
        variables = profile.get("variables", {})
        for col_name, stats in variables.items():
            pct_missing = stats.get("p_missing", 0)
            if pct_missing > 0.05:
                issues.append({
                    "field": col_name,
                    "issue": "high_null_rate",
                    "severity": "warning" if pct_missing < 0.3 else "error",
                    "detail": f"{pct_missing*100:.1f}% null values",
                    "action": "Review data source or add default value"
                })
            if stats.get("n_distinct", 0) == 1:
                issues.append({
                    "field": col_name,
                    "issue": "constant_value",
                    "severity": "info",
                    "detail": "Only 1 distinct value - may be redundant",
                    "action": "Consider removing or documenting as constant"
                })
        return issues

    for col_name, stats in columns.items():
        completeness = stats.get("completeness_pct", 100)
        if completeness < 95:
            issues.append({
                "field": col_name,
                "issue": "high_null_rate",
                "severity": "warning" if completeness > 70 else "error",
                "detail": f"{100-completeness:.1f}% null values",
                "action": "Review data source or add default value"
            })
        if stats.get("unique_count", 0) == 1 and stats.get("non_null_count", 0) > 0:
            issues.append({
                "field": col_name,
                "issue": "constant_value",
                "severity": "info",
                "detail": "Only 1 distinct value",
                "action": "Consider removing or documenting as constant"
            })
    return issues


def cross_reference_fields(
    product_fields: Dict[str, Dict],
    staging_mappings: List[Dict],
    normalize_fn=None
) -> List[Dict]:
    """
    Cross-reference data product fields against staging schema mappings.
    Uses data-cleaning-copilot's normalize_column_name when available.
    """
    gaps = []
    staged_fields = set()
    for mapping in staging_mappings:
        btp_field = mapping.get("btpField", "")
        staged_fields.add(btp_field.lower())
        if normalize_fn:
            staged_fields.add(normalize_fn(btp_field))

    for tech_name, field_info in product_fields.items():
        name_lower = tech_name.lower()
        normalized = normalize_fn(tech_name) if normalize_fn else name_lower
        if name_lower not in staged_fields and normalized not in staged_fields:
            gaps.append({
                "field": tech_name,
                "businessName": field_info.get("businessName", ""),
                "issue": "no_staging_mapping",
                "detail": "Field has no corresponding staging schema entry",
                "action": "Add to 2_stagingschema.csv or verify source system"
            })
    return gaps


def enrich_data_product(
    product_id: str,
    source_file: str,
    sheet_name: str = None,
    description: str = ""
) -> Dict[str, Any]:
    """
    Full enrichment pipeline for one data product.

    1. Load XLSX
    2. Profile with data-cleaning-copilot (or basic fallback)
    3. Classify fields
    4. Detect missing data
    5. Generate enrichment report
    """
    profile_fn, col_profile_fn = try_import_profiler()
    normalize_fn, mapping_fn, match_fn = try_import_column_mapping()

    filepath = DATA_DIR / source_file
    if not filepath.exists():
        return {"error": f"File not found: {filepath}"}

    headers, data = load_xlsx_sheet(str(filepath), sheet_name)

    # Profile
    if profile_fn:
        try:
            profile = pandas_profile(headers, data, profile_fn)
        except Exception as e:
            print(f"[WARN] Pandas profiling failed ({e}), using basic profiling")
            profile = basic_profile(headers, data)
    else:
        profile = basic_profile(headers, data)

    # Classify fields
    field_classifications = {}
    for i, col_name in enumerate(headers):
        values = [row[i] for row in data if i < len(row)]
        classification = classify_field(col_name, values, normalize_fn)
        field_classifications[col_name] = {
            **classification,
            "columnIndex": i,
        }

    # Detect issues
    quality_issues = detect_missing_data(profile)

    # Build enrichment report
    report = {
        "productId": product_id,
        "source": source_file,
        "sheet": sheet_name,
        "description": description,
        "enrichedAt": datetime.utcnow().isoformat() + "Z",
        "pipeline": "xlsx_to_odps.py",
        "copilotIntegration": {
            "profilerAvailable": profile_fn is not None,
            "columnMappingAvailable": normalize_fn is not None,
            "copilotPath": str(COPILOT_PATH)
        },
        "profile": {
            "rowCount": len(data),
            "columnCount": len(headers),
            "columns": headers
        },
        "fieldClassifications": field_classifications,
        "qualityIssues": quality_issues,
        "qualitySummary": {
            "totalFields": len(headers),
            "classifiedFields": sum(
                1 for f in field_classifications.values() if f["fieldType"]
            ),
            "measures": sum(
                1 for f in field_classifications.values()
                if f["fieldType"] == "measure"
            ),
            "dimensions": sum(
                1 for f in field_classifications.values()
                if f["fieldType"] == "dimension"
            ),
            "identifiers": sum(
                1 for f in field_classifications.values()
                if f["fieldType"] == "identifier"
            ),
            "dates": sum(
                1 for f in field_classifications.values()
                if f["fieldType"] == "date"
            ),
            "issueCount": len(quality_issues),
            "errorCount": sum(
                1 for i in quality_issues if i["severity"] == "error"
            ),
            "warningCount": sum(
                1 for i in quality_issues if i["severity"] == "warning"
            )
        }
    }

    return report


def load_staging_mappings() -> List[Dict]:
    """Load staging schema CSV for cross-referencing."""
    staging_file = DATA_DIR / "2_stagingschema.csv"
    if not staging_file.exists():
        return []
    mappings = []
    with open(staging_file, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            mappings.append(row)
    return mappings


def run_full_pipeline():
    """Run enrichment across all data products."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    products = [
        {
            "id": "treasury-capital-markets-v1",
            "source": "DATA_DICTIONARY.xlsx",
            "sheet": "Dictionary",
            "description": "Treasury Bond/Issuance positions"
        },
        {
            "id": "esg-net-zero-v1",
            "source": "ESG_ DATA_DICTIONARY.xlsx",
            "sheet": "NetZero",
            "description": "Net Zero emissions model"
        },
        {
            "id": "esg-client-v1",
            "source": "ESG_ DATA_DICTIONARY.xlsx",
            "sheet": "Client",
            "description": "Integrated Client ESG model"
        },
        {
            "id": "esg-sustainable-v1",
            "source": "ESG_ DATA_DICTIONARY.xlsx",
            "sheet": "Sustainable",
            "description": "Sustainable Finance model"
        },
        {
            "id": "performance-account-dim",
            "source": "NFRP_Account_AM.xlsx",
            "description": "Account dimension hierarchy"
        },
        {
            "id": "performance-location-dim",
            "source": "NFRP_Location_AM.xlsx",
            "description": "Location dimension hierarchy"
        },
        {
            "id": "performance-product-dim",
            "source": "NFRP_Product_AM.xlsx",
            "description": "Product dimension hierarchy"
        },
        {
            "id": "performance-segment-dim",
            "source": "NFRP_Segment_AM.xlsx",
            "description": "Segment dimension hierarchy"
        },
        {
            "id": "performance-cost-dim",
            "source": "NFRP_Cost_AM.xlsx",
            "description": "Cost centre dimension hierarchy"
        },
        {
            "id": "performance-crd-fact",
            "source": "Performance CRD - Fact table.xlsx",
            "description": "CRD Financial Performance fact table"
        },
    ]

    staging_mappings = load_staging_mappings()
    normalize_fn, _, _ = try_import_column_mapping()

    all_reports = []
    total_issues = 0
    total_errors = 0

    print("=" * 70)
    print("XLSX → ODPS 4.1 Enrichment Pipeline")
    print(f"Data dir: {DATA_DIR}")
    print(f"Copilot: {COPILOT_PATH}")
    print("=" * 70)

    for product in products:
        print(f"\n--- Processing: {product['id']} ---")
        report = enrich_data_product(
            product["id"],
            product["source"],
            product.get("sheet"),
            product.get("description", "")
        )

        if "error" in report:
            print(f"  ERROR: {report['error']}")
            continue

        summary = report["qualitySummary"]
        print(f"  Rows: {report['profile']['rowCount']}")
        print(f"  Columns: {summary['totalFields']}")
        print(f"  Classified: {summary['classifiedFields']} "
              f"(M:{summary['measures']} D:{summary['dimensions']} "
              f"I:{summary['identifiers']} T:{summary['dates']})")
        print(f"  Issues: {summary['issueCount']} "
              f"(errors:{summary['errorCount']}, warnings:{summary['warningCount']})")

        # Cross-reference against staging schema
        if staging_mappings and report.get("fieldClassifications"):
            gaps = cross_reference_fields(
                report["fieldClassifications"],
                staging_mappings,
                normalize_fn
            )
            report["stagingGaps"] = gaps
            if gaps:
                print(f"  Staging gaps: {len(gaps)} fields not in staging schema")

        # Write individual enrichment report
        outfile = OUTPUT_DIR / f"{product['id']}_enrichment.yaml"
        with open(outfile, "w") as f:
            yaml.dump(report, f, default_flow_style=False, sort_keys=False, allow_unicode=True)
        print(f"  → {outfile}")

        all_reports.append(report)
        total_issues += summary["issueCount"]
        total_errors += summary["errorCount"]

    # Write summary
    summary_report = {
        "pipeline": "xlsx_to_odps",
        "runAt": datetime.utcnow().isoformat() + "Z",
        "productsProcessed": len(all_reports),
        "totalIssues": total_issues,
        "totalErrors": total_errors,
        "products": [
            {
                "id": r["productId"],
                "rows": r["profile"]["rowCount"],
                "columns": r["profile"]["columnCount"],
                "issues": r["qualitySummary"]["issueCount"]
            }
            for r in all_reports
        ]
    }

    summary_file = OUTPUT_DIR / "pipeline_summary.yaml"
    with open(summary_file, "w") as f:
        yaml.dump(summary_report, f, default_flow_style=False, sort_keys=False)

    print(f"\n{'=' * 70}")
    print(f"DONE: {len(all_reports)} products enriched")
    print(f"Total issues: {total_issues} (errors: {total_errors})")
    print(f"Summary: {summary_file}")
    print(f"{'=' * 70}")

    return summary_report


if __name__ == "__main__":
    run_full_pipeline()
