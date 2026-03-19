#!/usr/bin/env python3
"""
Convert all .xlsx files in the training data directory to .csv for version control.
Also extracts real prompt samples from Excel files into JSONL training format.

Usage:
    python scripts/convert_xlsx_to_csv.py [--data-dir data/] [--output-dir data/csv/]
"""

import csv
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


DATA_DIR = Path(__file__).resolve().parent.parent / "data"
CSV_OUTPUT_DIR = DATA_DIR / "csv"
PROMPTS_OUTPUT = DATA_DIR / "massive_semantic" / "real_prompts.jsonl"


def xlsx_to_csv(xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Convert an xlsx file to one csv per sheet. Returns list of output paths."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    outputs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        safe_name = sheet_name.replace(" ", "_").replace("/", "_")
        stem = xlsx_path.stem.replace(" ", "_")
        out_path = output_dir / f"{stem}__{safe_name}.csv"
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in rows:
                writer.writerow([str(c) if c is not None else "" for c in row])
        outputs.append(out_path)
        print(f"  {sheet_name} -> {out_path.name} ({len(rows)} rows)")
    return outputs


def extract_prompt_samples() -> list[dict]:
    """Extract real prompts from the prompt sample xlsx files into training format."""
    examples = []

    # Treasury prompts (Prompt_samples.xlsx)
    treasury_path = DATA_DIR / "Prompt_samples.xlsx"
    if treasury_path.exists():
        wb = openpyxl.load_workbook(str(treasury_path), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:  # skip header
            if len(row) >= 4 and row[3]:
                examples.append({
                    "question": str(row[3]).strip(),
                    "template": str(row[2]).strip() if row[2] else "",
                    "category": str(row[0]).strip() if row[0] else "",
                    "product": str(row[1]).strip() if row[1] else "",
                    "domain": "treasury",
                    "type": "real_prompt",
                    "source": "Prompt_samples.xlsx",
                })

    # ESG prompts
    esg_path = DATA_DIR / "ESG_Prompt_samples.xlsx"
    if esg_path.exists():
        wb = openpyxl.load_workbook(str(esg_path), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if len(row) >= 3 and row[2]:
                examples.append({
                    "question": str(row[2]).strip(),
                    "template": str(row[1]).strip() if row[1] else "",
                    "product": str(row[0]).strip() if row[0] else "",
                    "domain": "esg",
                    "type": "real_prompt",
                    "source": "ESG_Prompt_samples.xlsx",
                })

    # Performance/BPC prompts
    bpc_path = DATA_DIR / "Performance (BPC) - sample prompts.xlsx"
    if bpc_path.exists():
        wb = openpyxl.load_workbook(str(bpc_path), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if len(row) >= 3 and row[2]:
                prompt_text = str(row[2]).strip()
                if prompt_text and prompt_text != "None":
                    metadata = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                    examples.append({
                        "question": prompt_text,
                        "metadata": metadata,
                        "domain": "performance",
                        "type": "real_prompt",
                        "source": "Performance_BPC_sample_prompts.xlsx",
                    })

    return examples


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Convert xlsx to csv and extract prompts")
    parser.add_argument("--data-dir", default=str(DATA_DIR))
    parser.add_argument("--output-dir", default=str(CSV_OUTPUT_DIR))
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Convert all xlsx files
    print("=== Converting XLSX to CSV ===")
    xlsx_files = sorted(data_dir.glob("*.xlsx"))
    total_csvs = 0
    for xlsx_file in xlsx_files:
        print(f"\n{xlsx_file.name}:")
        csvs = xlsx_to_csv(xlsx_file, output_dir)
        total_csvs += len(csvs)
    print(f"\nConverted {len(xlsx_files)} xlsx files -> {total_csvs} csv files in {output_dir}")

    # Extract real prompts
    print("\n=== Extracting Real Prompt Samples ===")
    prompts = extract_prompt_samples()
    PROMPTS_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(PROMPTS_OUTPUT, "w") as f:
        for p in prompts:
            f.write(json.dumps(p) + "\n")
    print(f"Extracted {len(prompts)} real prompts -> {PROMPTS_OUTPUT}")

    # Summary by domain
    from collections import Counter
    domains = Counter(p["domain"] for p in prompts)
    for d, c in domains.most_common():
        print(f"  {d}: {c} prompts")


if __name__ == "__main__":
    main()

